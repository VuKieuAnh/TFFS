from sklearn.model_selection import train_test_split # Import train_test_split function

from sklearn import metrics #Import scikit-learn metrics module for accuracy calculation
from sklearn.ensemble import RandomForestClassifier

import random
import numpy as np
import math
from sklearn import tree
from sklearn.naive_bayes import GaussianNB
from sklearn.naive_bayes import MultinomialNB, BernoulliNB
from sklearn.svm import SVC
from sklearn.naive_bayes import GaussianNB
from sklearn.naive_bayes import MultinomialNB, BernoulliNB
from sklearn.svm import SVC

def buildRFwithAllFeature(df, n):
    rf_model = RandomForestClassifier(n_estimators=500)
    df.columns.values[0] = "class"
    X = df.iloc[:, df.columns != 'class']
    Y = df[['class']]
    d = {}
    acc_RF = list()
    for i in range(n):
        X_Train, X_Test, Y_Train, Y_Test = train_test_split(X, Y, train_size=0.7,
                                                            random_state=np.random.randint(0, 100000))
        r, c = df.shape
        rf_model.fit(X_Train, Y_Train.values.ravel())
        pred_y = rf_model.predict(X_Test)
        accRFi = metrics.accuracy_score(Y_Test, pred_y)
        acc_RF.append(accRFi);
    print("Gia tri trung binh")
    print(np.mean(acc_RF))
    print("Do lech chuan")
    print(np.std(acc_RF))  ## Do lech chuan
    return acc_RF;


def buildRFwithArrFeature(df, n, arrFeature):
    rf_model = RandomForestClassifier(n_estimators=500)
    df.columns.values[0] = "class"
    X = df.iloc[:, df.columns != 'class']
    Y = df[['class']]
    acc_RF = list()
    for i in range(n):
        X_Train, X_Test, Y_Train, Y_Test = train_test_split(X, Y, train_size=0.7,
                                                            random_state=np.random.randint(0, 100000))
        X_Train1 = X_Train.iloc[:, arrFeature]
        X_Test1 = X_Test.iloc[:, arrFeature]
        rf_model.fit(X_Train1, Y_Train.values.ravel())
        pred_y = rf_model.predict(X_Test1)
        accRFi = metrics.accuracy_score(Y_Test, pred_y)
        acc_RF.append(accRFi)
    print("Gia tri trung binh")
    print(np.mean(acc_RF))  ## gia tri trung binh
    print("Do lech chuan")
    print(np.std(acc_RF))  ## Do lech chuan
    return acc_RF;


def buidSVMAllFeature(df, n):
    svclassifier = SVC(kernel='rbf')
    df.columns.values[0] = "class"
    X = df.iloc[:, df.columns != 'class']
    Y = df[['class']]
    acc_SVC = list()
    for i in range(n):
        X_Train, X_Test, Y_Train, Y_Test = train_test_split(X, Y, train_size=0.7,
                                                            random_state=np.random.randint(0, 100000))
        svclassifier.fit(X_Train, Y_Train.values.ravel())
        y_pred = svclassifier.predict(X_Test)
        accI = metrics.accuracy_score(Y_Test, y_pred);
        acc_SVC.append(accI);
    print("Gia tri trung binh")
    print(np.mean(acc_SVC))  ## gia tri trung binh
    print("Do lech chuan")
    print(np.std(acc_SVC))  ## Do lech chuan
    return acc_SVC


def buidSVMSSelectedFeature(df, n, arrFeature):
    svclassifier = SVC(kernel='rbf')
    df.columns.values[0] = "class"
    X = df.iloc[:, df.columns != 'class']
    Y = df[['class']]
    d = {}
    acc_SVC = list()
    for i in range(n):
        X_Train, X_Test, Y_Train, Y_Test = train_test_split(X, Y, train_size=0.7,
                                                            random_state=np.random.randint(0, 100000))
        X_Train1 = X_Train.iloc[:, arrFeature]
        X_Test1 = X_Test.iloc[:, arrFeature]
        svclassifier.fit(X_Train1, Y_Train.values.ravel())
        y_pred = svclassifier.predict(X_Test1)
        accI = metrics.accuracy_score(Y_Test, y_pred);
        acc_SVC.append(accI);
    print("Gia tri trung binh")
    print(np.mean(acc_SVC))  ## gia tri trung binh
    print("Do lech chuan")
    print(np.std(acc_SVC))  ## Do lech chuan
    return acc_SVC


def buidNaivebayesAllFeature(df, n):
    model_navie = GaussianNB()
    df.columns.values[0] = "class"
    X = df.iloc[:, df.columns != 'class']
    Y = df[['class']]
    d = {}
    acc_NB = list()
    for i in range(n):
        X_Train, X_Test, Y_Train, Y_Test = train_test_split(X, Y, train_size=0.7,
                                                            random_state=np.random.randint(0, 100000))
        model_navie.fit(X_Train, Y_Train.values.ravel())
        prediction = model_navie.predict(X_Test)
        accI = metrics.accuracy_score(Y_Test, prediction);
        acc_NB.append(accI);
    print("Gia tri trung binh")
    print(np.mean(acc_NB))  ## gia tri trung binh
    print("Do lech chuan")
    print(np.std(acc_NB))  ## Do lech chuan
    return acc_NB


def buidNaivebayesSelectedFeature(df, n, arrFeature):
    model_navie = GaussianNB()
    df.columns.values[0] = "class"
    X = df.iloc[:, df.columns != 'class']
    Y = df[['class']]
    d = {}
    acc_NB = list()
    for i in range(n):
        X_Train, X_Test, Y_Train, Y_Test = train_test_split(X, Y, train_size=0.7,
                                                            random_state=np.random.randint(0, 100000))
        X_Train1 = X_Train.iloc[:, arrFeature]
        X_Test1 = X_Test.iloc[:, arrFeature]
        model_navie.fit(X_Train1, Y_Train.values.ravel())
        prediction = model_navie.predict(X_Test1)
        accI = metrics.accuracy_score(Y_Test, prediction);
        acc_NB.append(accI);
    print("Gia tri trung binh")
    print(np.mean(acc_NB))  ## gia tri trung binh
    print("Do lech chuan")
    print(np.std(acc_NB))  ## Do lech chuan
    return acc_NB


import openpyxl

def output_Excel(input_detail, output_excel_path, fileCSV):
    # Xác định số hàng và cột lớn nhất trong file excel cần tạo
    row = 21
    column = 6

    # Tạo một workbook mới và active nó
    wb = openpyxl.load_workbook(output_excel_path)
    wb.create_sheet(fileCSV)
    sheet = wb[fileCSV]
    #   ws = wb.active

    # Dùng vòng lặp for để ghi nội dung từ input_detail vào file Excel
    for i in range(0, column):
        for j in range(0, row):
            v = input_detail[i][j]
            sheet.cell(row=j + 1, column=i + 1, value=v)

    # Lưu lại file Excel
    wb.save(output_excel_path)
